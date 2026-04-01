"""
╔══════════════════════════════════════════════════════════════════╗
║       QR CODE SECURITY SCANNER — SEMI-AUTOMATED TESTER          ║
║  Cara pakai:                                                     ║
║  1. pip install qrcode pillow openpyxl                          ║
║  2. Siapkan dataset CSV (kolom: url, label, kategori)           ║
║  3. python qr_scanner_tester.py                                  ║
╚══════════════════════════════════════════════════════════════════╝
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import qrcode
from PIL import Image, ImageTk, ImageDraw, ImageFont
import csv
import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Konfigurasi Scanner yang Diuji ───────────────────────────────────────
SCANNERS = [
    "iOS Camera (Built-in)",
    "Samsung Camera (Built-in)",
    "Redmi Camera (Built-in)",
    "Google Lens",
    "Lionic Qr Scanner",
    "QR Scan Shield"   # ganti sesuai kebutuhan
]

# ─── Konfigurasi Warna ────────────────────────────────────────────────────
BG_DARK   = "#1a1a2e"
BG_CARD   = "#16213e"
BG_PANEL  = "#0f3460"
ACCENT    = "#e94560"
GREEN_OK  = "#27ae60"
RED_WARN  = "#e74c3c"
ORANGE    = "#f39c12"
BLUE_BTN  = "#2980b9"
TEXT_MAIN = "#ecf0f1"
TEXT_DIM  = "#95a5a6"

# ─── Util: border openpyxl ────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

# ─── QR Code Generator ────────────────────────────────────────────────────
def generate_qr(url: str, size: int = 300) -> Image.Image:
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    img = img.resize((size, size), Image.LANCZOS)
    return img

# ─── Excel Logger ─────────────────────────────────────────────────────────
class ExcelLogger:
    def __init__(self, path: str):
        self.path = path
        self._init_workbook()

    def _init_workbook(self):
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
        else:
            self.wb = Workbook()
            self._setup_sheets()
            self.wb.save(self.path)

    def _hcell(self, ws, r, c, val, bg="1F3864", fg="FFFFFF", bold=True, size=10):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        return cell

    def _dcell(self, ws, r, c, val, bg="FFFFFF", fg="000000", bold=False, fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(color=fg, size=10, name="Arial", bold=bold)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        if fmt:
            cell.number_format = fmt
        return cell

    def _setup_sheets(self):
        wb = self.wb
        # Sheet per scanner
        for scanner in SCANNERS:
            safe_name = scanner[:31]
            if safe_name not in wb.sheetnames:
                ws = wb.create_sheet(safe_name)
                self._setup_scanner_sheet(ws, scanner)

        # Sheet Ringkasan
        if "Ringkasan" not in wb.sheetnames:
            ws_sum = wb.create_sheet("Ringkasan")
            self._setup_summary_sheet(ws_sum)

        # Hapus sheet default kosong
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    def _setup_scanner_sheet(self, ws, scanner_name):
        ws.merge_cells("A1:L1")
        t = ws["A1"]
        t.value = f"LOG PENGUJIAN: {scanner_name.upper()}"
        t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
        t.fill = PatternFill("solid", start_color="1F3864")
        t.alignment = Alignment(horizontal="center", vertical="center")
        t.border = thin_border()
        ws.row_dimensions[1].height = 28

        headers = [
            ("A", 5,  "No"),
            ("B", 12, "Waktu"),
            ("C", 40, "URL"),
            ("D", 12, "Label\nAktual"),
            ("E", 15, "Kategori"),
            ("F", 14, "Hasil\nScanner"),
            ("G", 10, "TP"),
            ("H", 10, "FP"),
            ("I", 10, "FN"),
            ("J", 10, "TN"),
            ("K", 14, "Benar?"),
            ("L", 25, "Catatan"),
        ]
        col_colors = {
            "G": "2E75B6", "H": "C00000",
            "I": "C55A11", "J": "375623",
        }
        for col_l, width, label in headers:
            ws.column_dimensions[col_l].width = width
            bg = col_colors.get(col_l, "1F3864")
            col_idx = ord(col_l) - ord("A") + 1
            self._hcell(ws, 2, col_idx, label, bg=bg, size=9)
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

    def _setup_summary_sheet(self, ws):
        ws.merge_cells("A1:H1")
        t = ws["A1"]
        t.value = "RINGKASAN PERBANDINGAN SEMUA SCANNER"
        t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
        t.fill = PatternFill("solid", start_color="1F3864")
        t.alignment = Alignment(horizontal="center", vertical="center")
        t.border = thin_border()
        ws.row_dimensions[1].height = 28

        for w, col in [("A",30),("B",10),("C",10),("D",10),
                       ("E",10),("F",12),("G",12),("H",12)]:
            ws.column_dimensions[w].width = col

        hdrs = ["Scanner", "TP", "FP", "FN", "TN", "Precision", "Recall", "F1-Score"]
        for ci, h in enumerate(hdrs, 1):
            self._hcell(ws, 2, ci, h, bg="2E75B6")

        for ri, scanner in enumerate(SCANNERS, 3):
            safe = scanner[:31]
            ws_ref = f"'{safe}'"
            ws.row_dimensions[ri].height = 20
            bg = "DEEAF1" if ri % 2 == 0 else "FFFFFF"

            self._dcell(ws, ri, 1, scanner, bg=bg, bold=True)

            tp_f = f"=IFERROR(SUMIF({ws_ref}!D$3:D$1000,\"phishing\",{ws_ref}!G$3:G$1000),0)"
            fp_f = f"=IFERROR(SUMIF({ws_ref}!D$3:D$1000,\"safe\",{ws_ref}!H$3:H$1000),0)"
            fn_f = f"=IFERROR(SUMIF({ws_ref}!D$3:D$1000,\"phishing\",{ws_ref}!I$3:I$1000),0)"
            tn_f = f"=IFERROR(SUMIF({ws_ref}!D$3:D$1000,\"safe\",{ws_ref}!J$3:J$1000),0)"

            for ci, formula in enumerate([tp_f, fp_f, fn_f, tn_f], 2):
                self._dcell(ws, ri, ci, formula, bg=bg, bold=True)

            col_b = chr(ord('A') + 1)  # B
            col_c = chr(ord('A') + 2)  # C
            col_d = chr(ord('A') + 3)  # D
            col_e = chr(ord('A') + 4)  # E

            prec_f = f"=IFERROR(B{ri}/(B{ri}+C{ri}),0)"
            rec_f  = f"=IFERROR(B{ri}/(B{ri}+D{ri}),0)"
            f1_f   = f"=IFERROR(2*F{ri}*G{ri}/(F{ri}+G{ri}),0)"

            self._dcell(ws, ri, 6, prec_f, bg=bg, fmt="0.00%")
            self._dcell(ws, ri, 7, rec_f,  bg=bg, fmt="0.00%")
            self._dcell(ws, ri, 8, f1_f,   bg=bg, fmt="0.00%")

    def log_result(self, scanner: str, url: str, label: str,
                   kategori: str, hasil_scanner: str, catatan: str = ""):
        """
        label         : 'phishing' atau 'safe'
        hasil_scanner : 'detected' (scanner bilang phishing) atau 'safe' (scanner bilang aman)
        """
        safe_name = scanner[:31]
        if safe_name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(safe_name)
            self._setup_scanner_sheet(ws, scanner)
        ws = self.wb[safe_name]

        # Cari baris berikutnya
        row = ws.max_row + 1
        if row < 3:
            row = 3

        no        = row - 2
        waktu     = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        hasil_str = "Phishing" if hasil_scanner == "detected" else "Aman"

        # Confusion matrix
        tp = fp = fn = tn = 0
        if label == "phishing" and hasil_scanner == "detected":
            tp = 1; benar = "✅"
        elif label == "safe" and hasil_scanner == "detected":
            fp = 1; benar = "❌"
        elif label == "phishing" and hasil_scanner == "safe":
            fn = 1; benar = "❌"
        else:
            tn = 1; benar = "✅"

        bg = "E2EFDA" if benar == "✅" else "FFE0E0"

        data = [no, waktu, url, label, kategori, hasil_str, tp, fp, fn, tn, benar, catatan]
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = Font(size=9, name="Arial")
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center" if ci != 3 else "left",
                                       vertical="center")
            cell.border = thin_border()

        self.wb.save(self.path)


# ─── Main GUI Application ─────────────────────────────────────────────────
class ScannerTesterApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("QR Security Scanner Tester")
        self.root.geometry("900x700")
        self.root.configure(bg=BG_DARK)
        self.root.resizable(True, True)

        self.urls        = []          # list of dict {url, label, kategori}
        self.idx         = 0           # indeks URL saat ini
        self.scanner_idx = 0           # indeks scanner saat ini
        self.excel_path  = ""
        self.logger      = None
        self.qr_image    = None
        self.scan_count  = {s: {"tp":0,"fp":0,"fn":0,"tn":0} for s in SCANNERS}
        self.session_results = []

        self._build_ui()

    # ── UI Builder ────────────────────────────────────────────────────────
    def _build_ui(self):
        # ─ Top Bar
        top = tk.Frame(self.root, bg=BG_PANEL, height=60)
        top.pack(fill="x")
        tk.Label(top, text="🔍 QR Security Scanner Tester",
                 bg=BG_PANEL, fg=TEXT_MAIN,
                 font=("Arial", 16, "bold")).pack(side="left", padx=20, pady=15)

        btn_frame = tk.Frame(top, bg=BG_PANEL)
        btn_frame.pack(side="right", padx=10, pady=10)
        tk.Button(btn_frame, text="📂 Load Dataset CSV",
                  command=self._load_csv,
                  bg=BLUE_BTN, fg="white", font=("Arial", 10, "bold"),
                  relief="flat", padx=12, pady=6, cursor="hand2").pack(side="left", padx=5)
        
        tk.Button(btn_frame, text="📥 Buka Excel Lama",
                  command=self._open_existing_excel,
                  bg=GREEN_OK, fg="white", font=("Arial", 10, "bold"),
                  relief="flat", padx=12, pady=6, cursor="hand2").pack(side="left", padx=5)
                  
        tk.Button(btn_frame, text="📄 Buat Excel Baru",
                  command=self._create_new_excel,
                  bg=ORANGE, fg="white", font=("Arial", 10, "bold"),
                  relief="flat", padx=12, pady=6, cursor="hand2").pack(side="left", padx=5)

        # ─ Status bar
        self.status_var = tk.StringVar(value="⚠ Belum ada dataset. Load CSV terlebih dahulu.")
        tk.Label(self.root, textvariable=self.status_var,
                 bg=BG_DARK, fg=ORANGE,
                 font=("Arial", 10, "italic")).pack(fill="x", padx=20, pady=4)

        # ─ Main content
        main = tk.Frame(self.root, bg=BG_DARK)
        main.pack(fill="both", expand=True, padx=20, pady=5)

        # Left: QR Display
        left = tk.Frame(main, bg=BG_CARD, bd=0, relief="flat")
        left.pack(side="left", fill="both", expand=True, padx=(0,10))

        tk.Label(left, text="QR CODE", bg=BG_CARD, fg=TEXT_DIM,
                 font=("Arial", 9, "bold")).pack(pady=(15,5))

        self.qr_label = tk.Label(left, bg=BG_CARD, text="",
                                 relief="flat")
        self.qr_label.pack(pady=5)

        self.url_var = tk.StringVar(value="—")
        tk.Label(left, textvariable=self.url_var, bg=BG_CARD, fg=TEXT_MAIN,
                 font=("Arial", 9), wraplength=300).pack(pady=5, padx=10)

        label_frame = tk.Frame(left, bg=BG_CARD)
        label_frame.pack(pady=5)
        tk.Label(label_frame, text="Label Aktual:", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Arial", 9)).pack(side="left")
        self.label_badge = tk.Label(label_frame, text="—", bg=BG_CARD,
                                    font=("Arial", 10, "bold"))
        self.label_badge.pack(side="left", padx=5)

        # Counter
        self.counter_var = tk.StringVar(value="0 / 0")
        tk.Label(left, textvariable=self.counter_var, bg=BG_CARD,
                 fg=TEXT_DIM, font=("Arial", 9)).pack(pady=(2,10))

        # Right: Controls
        right = tk.Frame(main, bg=BG_CARD, width=280)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        # Scanner selector
        tk.Label(right, text="SCANNER YANG DIUJI", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Arial", 9, "bold")).pack(pady=(18,5))

        self.scanner_var = tk.StringVar(value=SCANNERS[0])
        scanner_menu = ttk.Combobox(right, textvariable=self.scanner_var,
                                    values=SCANNERS, state="readonly",
                                    font=("Arial", 10), width=26)
        scanner_menu.pack(padx=15, pady=3)

        # Navigation
        tk.Label(right, text="NAVIGASI URL", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Arial", 9, "bold")).pack(pady=(15,5))

        nav_frame = tk.Frame(right, bg=BG_CARD)
        nav_frame.pack(pady=3)
        tk.Button(nav_frame, text="◀ Prev",
                  command=self._prev_url,
                  bg=BG_PANEL, fg=TEXT_MAIN, font=("Arial", 9),
                  relief="flat", padx=10, pady=5, cursor="hand2").pack(side="left", padx=5)
        tk.Button(nav_frame, text="Next ▶",
                  command=self._next_url,
                  bg=BG_PANEL, fg=TEXT_MAIN, font=("Arial", 9),
                  relief="flat", padx=10, pady=5, cursor="hand2").pack(side="left", padx=5)

        # ── HASIL SCAN INPUT ──
        tk.Label(right, text="HASIL SCAN DI HP", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Arial", 9, "bold")).pack(pady=(18,8))

        tk.Label(right, text="Scanner HP mendeteksi URL ini sebagai:",
                 bg=BG_CARD, fg=TEXT_MAIN,
                 font=("Arial", 9), wraplength=240).pack()

        # Tombol besar
        self.btn_detected = tk.Button(
            right,
            text="🚨  PHISHING TERDETEKSI\n(Scanner memberi peringatan)",
            command=lambda: self._record("detected"),
            bg=RED_WARN, fg="white", font=("Arial", 10, "bold"),
            relief="flat", padx=10, pady=12,
            cursor="hand2", wraplength=220
        )
        self.btn_detected.pack(padx=15, pady=(8,4), fill="x")

        self.btn_safe = tk.Button(
            right,
            text="✅  AMAN / TIDAK ADA PERINGATAN\n(Scanner membuka langsung)",
            command=lambda: self._record("safe"),
            bg=GREEN_OK, fg="white", font=("Arial", 10, "bold"),
            relief="flat", padx=10, pady=12,
            cursor="hand2", wraplength=220
        )
        self.btn_safe.pack(padx=15, pady=4, fill="x")

        self.btn_noresult = tk.Button(
            right,
            text="⚠  TIDAK BISA SCAN\n(QR tidak terbaca)",
            command=lambda: self._record("noscan"),
            bg=ORANGE, fg="white", font=("Arial", 10, "bold"),
            relief="flat", padx=10, pady=8,
            cursor="hand2", wraplength=220
        )
        self.btn_noresult.pack(padx=15, pady=4, fill="x")

        # Catatan
        tk.Label(right, text="Catatan (opsional):", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Arial", 9)).pack(pady=(12,3))
        self.note_entry = tk.Entry(right, font=("Arial", 9),
                                   bg=BG_PANEL, fg=TEXT_MAIN,
                                   insertbackground=TEXT_MAIN,
                                   relief="flat", width=28)
        self.note_entry.pack(padx=15, pady=3, ipady=5)

        # Mini scoreboard
        tk.Label(right, text="SKOR SESI INI", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Arial", 9, "bold")).pack(pady=(15,5))
        self.score_var = tk.StringVar(value="TP:0  FP:0  FN:0  TN:0")
        tk.Label(right, textvariable=self.score_var,
                 bg=BG_CARD, fg=ACCENT,
                 font=("Arial", 11, "bold")).pack()

        # Keyboard shortcuts
        tk.Label(right, text="Shortcut: [1]=Phishing  [2]=Aman  [3]=NoScan",
                 bg=BG_CARD, fg=TEXT_DIM,
                 font=("Arial", 8)).pack(pady=(8,0))

        self.root.bind("1", lambda e: self._record("detected"))
        self.root.bind("2", lambda e: self._record("safe"))
        self.root.bind("3", lambda e: self._record("noscan"))
        self.root.bind("<Right>", lambda e: self._next_url())
        self.root.bind("<Left>",  lambda e: self._prev_url())

        self._disable_scan_buttons()

    # ── Actions ───────────────────────────────────────────────────────────
    def _load_csv(self):
        path = filedialog.askopenfilename(
            title="Pilih Dataset CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return
        self.urls = []
        with open(path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                self.urls.append({
                    "url":      row.get("url", "").strip(),
                    "label":    row.get("label", "safe").strip().lower(),
                    "kategori": row.get("kategori", "").strip(),
                })
        self.idx = 0
        total   = len(self.urls)
        phish   = sum(1 for u in self.urls if u["label"] == "phishing")
        safe    = total - phish
        self.status_var.set(
            f"✅ Dataset dimuat: {total} URL  |  🔴 Phishing: {phish}  |  🟢 Safe: {safe}"
        )
        self._show_current()
        self._enable_scan_buttons()
        messagebox.showinfo("Dataset Dimuat",
                            f"{total} URL berhasil dimuat!\nPhishing: {phish}  |  Safe: {safe}")

    def _open_existing_excel(self):
        """Fungsi untuk memilih file Excel yang sudah berisi data (punya teman)"""
        path = filedialog.askopenfilename(
            title="Pilih File Excel Lama",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return
        self._set_excel_logger(path, is_new=False)

    def _create_new_excel(self):
        """Fungsi untuk membuat file Excel dari nol (kosong)"""
        path = filedialog.asksaveasfilename(
            title="Buat File Excel Baru",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="QR_Scanner_Test_Log.xlsx"
        )
        if not path:
            return
        self._set_excel_logger(path, is_new=True)

    def _set_excel_logger(self, path, is_new):
        """Fungsi bantuan untuk setup logger dan update UI"""
        self.excel_path = path
        self.logger = ExcelLogger(path)
        
        status_teks = "Dibuat" if is_new else "Dibuka (Melanjutkan Data)"
        self.status_var.set(f"💾 Output Excel {status_teks}: {os.path.basename(path)}")
        messagebox.showinfo("Excel Siap", f"Log akan disimpan ke:\n{path}")

    def _show_current(self):
        if not self.urls:
            return
        entry = self.urls[self.idx]
        url   = entry["url"]
        label = entry["label"]

        # Generate QR
        img = generate_qr(url, size=260)
        self.qr_image = ImageTk.PhotoImage(img)
        self.qr_label.configure(image=self.qr_image)

        # Update text
        display_url = url if len(url) <= 55 else url[:52] + "..."
        self.url_var.set(display_url)

        color = RED_WARN if label == "phishing" else GREEN_OK
        badge = "🔴 PHISHING" if label == "phishing" else "🟢 SAFE"
        self.label_badge.configure(text=badge, fg=color, bg=BG_CARD)

        self.counter_var.set(f"{self.idx + 1} / {len(self.urls)}")

    def _next_url(self):
        if self.urls and self.idx < len(self.urls) - 1:
            self.idx += 1
            self._show_current()

    def _prev_url(self):
        if self.urls and self.idx > 0:
            self.idx -= 1
            self._show_current()

    def _record(self, hasil: str):
        if not self.urls:
            messagebox.showwarning("Belum ada data", "Load dataset CSV terlebih dahulu!")
            return
        if not self.excel_path:
            messagebox.showwarning("Belum ada output", "Pilih file Excel output terlebih dahulu!")
            return

        if hasil == "noscan":
            messagebox.showinfo("Dicatat", "URL dilewati (tidak dapat di-scan). Lanjut ke berikutnya.")
            self._next_url()
            return

        entry   = self.urls[self.idx]
        scanner = self.scanner_var.get()
        catatan = self.note_entry.get()
        self.note_entry.delete(0, "end")

        # Hitung confusion
        label = entry["label"]
        tp = fp = fn = tn = 0
        if label == "phishing" and hasil == "detected":
            tp = 1
        elif label == "safe" and hasil == "detected":
            fp = 1
        elif label == "phishing" and hasil == "safe":
            fn = 1
        else:
            tn = 1

        sc = self.scan_count[scanner]
        sc["tp"] += tp; sc["fp"] += fp
        sc["fn"] += fn; sc["tn"] += tn

        self.score_var.set(
            f"TP:{sc['tp']}  FP:{sc['fp']}  FN:{sc['fn']}  TN:{sc['tn']}"
        )

        # Log ke Excel
        self.logger.log_result(
            scanner  = scanner,
            url      = entry["url"],
            label    = label,
            kategori = entry["kategori"],
            hasil_scanner = hasil,
            catatan  = catatan,
        )

        # Feedback visual
        if tp or tn:
            flash = GREEN_OK
        else:
            flash = RED_WARN
        self.root.configure(bg=flash)
        self.root.after(200, lambda: self.root.configure(bg=BG_DARK))

        # Auto next
        self._next_url()

    def _disable_scan_buttons(self):
        self.btn_detected.configure(state="disabled")
        self.btn_safe.configure(state="disabled")
        self.btn_noresult.configure(state="disabled")

    def _enable_scan_buttons(self):
        self.btn_detected.configure(state="normal")
        self.btn_safe.configure(state="normal")
        self.btn_noresult.configure(state="normal")


# ─── Entry Point ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    app  = ScannerTesterApp(root)
    root.mainloop()
