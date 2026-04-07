"""
╔══════════════════════════════════════════════════════════════════╗
║      ROBOT GOOGLE LENS (DEEP CHECK - SATU JENDELA ANTI ERROR)    ║
╚══════════════════════════════════════════════════════════════════╝
"""

import os
import time
import csv
import datetime
import re
import random
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── BANTUAN EXCEL LOGGER ────
def thin_border():
    return Border(left=Side(style="thin"), right=Side(style="thin"), 
                  top=Side(style="thin"), bottom=Side(style="thin"))

def dcell(ws, r, c, val, bg="FFFFFF", fg="000000", bold=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(color=fg, size=10, name="Arial", bold=bold)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    if fmt: cell.number_format = fmt
    return cell

class AutoExcelLogger:
    def __init__(self, path="QR_Scanner_Test_Log.xlsx"):
        self.path = path
        self.scanner_name = "Google Lens (Auto)"
        
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
        else:
            from openpyxl import Workbook
            self.wb = Workbook()
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]
            
        self._setup_data_sheet()
        self._setup_summary_sheet()

    def _setup_data_sheet(self):
        safe_name = self.scanner_name[:31]
        if safe_name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(safe_name, 0)
            headers = ["No", "Waktu", "URL", "Label Aktual", "Kategori", "Hasil Scanner", "TP", "FP", "FN", "TN", "Benar?", "Catatan Lens"]
            for col_idx, label in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=label)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", start_color="1F3864")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border()
            ws.freeze_panes = "A3"
            self.wb.save(self.path)

    def _setup_summary_sheet(self):
        if "Ringkasan" not in self.wb.sheetnames:
            ws = self.wb.create_sheet("Ringkasan")
        else:
            ws = self.wb["Ringkasan"]

        ws.merge_cells("A1:H1")
        t = ws["A1"]
        t.value = "RINGKASAN METRIK OTOMASI"
        t.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
        t.fill = PatternFill("solid", start_color="1F3864")
        t.alignment = Alignment(horizontal="center", vertical="center")

        hdrs = ["Scanner", "TP", "FP", "FN", "TN", "Precision", "Recall", "F1-Score"]
        for ci, h in enumerate(hdrs, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="2E75B6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

        ri = 3
        bg = "DEEAF1"
        ws_ref = f"'{self.scanner_name[:31]}'"
        
        dcell(ws, ri, 1, self.scanner_name, bg=bg, bold=True)

        tp_f = f'=IFERROR(SUM({ws_ref}!G$3:G$5000),0)'
        fp_f = f'=IFERROR(SUM({ws_ref}!H$3:H$5000),0)'
        fn_f = f'=IFERROR(SUM({ws_ref}!I$3:I$5000),0)'
        tn_f = f'=IFERROR(SUM({ws_ref}!J$3:J$5000),0)'

        for ci, formula in enumerate([tp_f, fp_f, fn_f, tn_f], 2):
            dcell(ws, ri, ci, formula, bg=bg, bold=True)

        prec_f = f"=IFERROR(B{ri}/(B{ri}+C{ri}),0)"
        rec_f  = f"=IFERROR(B{ri}/(B{ri}+D{ri}),0)"
        f1_f   = f"=IFERROR(2*F{ri}*G{ri}/(F{ri}+G{ri}),0)"

        dcell(ws, ri, 6, prec_f, bg=bg, fmt="0.00%")
        dcell(ws, ri, 7, rec_f,  bg=bg, fmt="0.00%")
        dcell(ws, ri, 8, f1_f,   bg=bg, fmt="0.00%")
        
        for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions["A"].width = 25

        self.wb.save(self.path)

    def log_result(self, url, label, kategori, hasil_scanner, catatan=""):
        safe_name = self.scanner_name[:31]
        ws = self.wb[safe_name]
        row = ws.max_row + 1
        if row < 3: row = 3

        no = row - 2
        waktu = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        tp = fp = fn = tn = 0
        benar = "❓"
        
        if hasil_scanner == "error":
            hasil_str = "ERROR (Gagal Scan)"
            benar = "➖" 
            bg = "F2F2F2" 
        else:
            hasil_str = "Phishing" if hasil_scanner == "detected" else "Aman"
            if label == "phishing" and hasil_scanner == "detected": tp = 1; benar = "✅"
            elif label == "safe" and hasil_scanner == "detected": fp = 1; benar = "❌"
            elif label == "phishing" and hasil_scanner == "safe": fn = 1; benar = "❌"
            else: tn = 1; benar = "✅"
            bg = "E2EFDA" if benar == "✅" else "FFE0E0"

        data = [no, waktu, url, label, kategori, hasil_str, tp, fp, fn, tn, benar, catatan]
        
        for ci, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci != 3 else "left")

        self.wb.save(self.path)

# ─── FUNGSI BANTUAN ───
def jeda_manusia(min_detik=2.0, max_detik=4.5):
    time.sleep(random.uniform(min_detik, max_detik))

# ─── ROBOT OTOMASI SELENIUM ───────────────────────────────────────
def jalankan_robot():
    file_csv = "dataset(2).csv"
    folder_gambar = "qr_test_cases(2)"
    
    if not os.path.exists(file_csv):
        print("❌ File dataset.csv tidak ditemukan!")
        return

    logger = AutoExcelLogger()
    print("✅ Excel Logger & Sheet Ringkasan Metrik siap! (DEEP CHECK MODE)")

    print("🤖 Menyalakan mesin Chrome Penyamaran...")
    options = uc.ChromeOptions()
    options.add_argument('--lang=en')
    driver = uc.Chrome(options=options, version_main=146)
    driver.maximize_window()

    with open(file_csv, mode='r', encoding='utf-8') as f:
        reader = list(csv.DictReader(f))
        
    print(f"📂 Ditemukan {len(reader)} data untuk diuji.")

    try:
        for ri, row in enumerate(reader, 1):
            url_asli = row['url'].strip()
            label = row['label'].strip()
            kategori = row['kategori'].strip()
            
            file_safe_url = re.sub(r'[^a-zA-Z0-9]', '_', url_asli[:20])
            nama_file = f"qr_{ri:03}_{file_safe_url}.png"
            path_lengkap = os.path.abspath(os.path.join(folder_gambar, nama_file))
            
            if not os.path.exists(path_lengkap):
                continue

            print(f"\n🔍 [Scan {ri}/{len(reader)}] Sedang menguji: {url_asli}")
            
            upload_sukses = False

            # FASE 1: UPLOAD GAMBAR
            for attempt in range(2):
                step_sekarang = "Membuka Google Images"
                try:
                    driver.get('https://images.google.com/')
                    jeda_manusia(1.5, 2.5)
                    
                    while "sorry" in driver.current_url or "captcha" in driver.page_source.lower():
                        print("\n🚨🚨 CAPTCHA TERDETEKSI! 🚨🚨")
                        input("👉 Tolong selesaikan di Chrome, lalu tekan ENTER di sini...")
                        driver.get('https://images.google.com/')
                        jeda_manusia(2.0, 3.0)

                    try:
                        driver.execute_script("""
                            document.querySelectorAll('iframe').forEach(e => e.remove());
                            document.querySelectorAll('[role="dialog"]').forEach(e => e.remove());
                        """)
                        tombol_setuju = driver.find_elements(By.XPATH, "//button[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'setuju') or contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'accept')]")
                        for btn in tombol_setuju:
                            if btn.is_displayed():
                                driver.execute_script("arguments[0].click();", btn)
                    except:
                        pass 

                    kamera = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, "//*[contains(@aria-label, 'image') or contains(@aria-label, 'gambar') or contains(@data-tooltip, 'image')]"))
                    )
                    
                    try:
                        ActionChains(driver).move_to_element(kamera).click().perform()
                    except:
                        driver.execute_script("arguments[0].dispatchEvent(new MouseEvent('click', {bubbles: true}));", kamera)
                        
                    jeda_manusia(1.0, 2.0)
                    
                    upload_input = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
                    )
                    upload_input.send_keys(path_lengkap)
                    upload_sukses = True 
                    break 
                    
                except Exception as e:
                    if attempt == 0:
                        print(f"   ⚠️ Terhalang saat '{step_sekarang}'. Refresh dan coba lagi...")
                    else:
                        pesan_error = str(e).split('\n')[0]
                        if "Timeout" in pesan_error or "timeout" in str(type(e)).lower():
                            pesan_error = "Timeout/Terhalang Pop-up"
                        print(f"   🔴 Hasil Lens: DIBLOKIR / GAGAL AKSES ({pesan_error})")
                        logger.log_result(url_asli, label, kategori, "detected", f"Diblokir (Error UI: {pesan_error})")

            if not upload_sukses:
                continue 

            # FASE 2: SUPER EKSTRAKTOR & DEEP CHECK
            try:
                jeda_manusia(6.0, 8.0) 
                
                body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
                
                semua_link_element = driver.find_elements(By.TAG_NAME, "a")
                semua_href = [str(link.get_attribute('href')).lower() for link in semua_link_element if link.get_attribute('href')]

                url_mentah = url_asli.replace("https://", "").replace("http://", "").replace("www.", "").lower()
                domain_utama = url_mentah.split('/')[0]

                potongan_kata = [p for p in re.split(r'[./\-_]', url_mentah) if len(p) >= 4]

                ditemukan_di_teks = False
                if domain_utama in body_text:
                    ditemukan_di_teks = True
                else:
                    for part in potongan_kata:
                        if part in body_text:
                            ditemukan_di_teks = True
                            break
                            
                indikator_tombol = ["website", "situs web", "visit"]
                ada_tombol = any(kata in body_text for kata in indikator_tombol)
                ada_link_hidup = any(domain_utama in href for href in semua_href)

                if ada_link_hidup or (ditemukan_di_teks and ada_tombol):
                    # 🌟 FASE 3: DEEP CHECK (TIDAK BUKA TAB BARU, LANGSUNG HAJAR DI TAB YANG SAMA)
                    print("   🔍 Lens menemukan link. Melakukan navigasi ke URL untuk verifikasi keamanan...")
                    
                    try:
                        # Waktu maksimal buat loading page adalah 15 detik. Kalau lebih, anggap server mati/diblokir.
                        driver.set_page_load_timeout(15) 
                        driver.get(url_asli) # Langsung pindah halaman!
                        jeda_manusia(3.0, 5.0)
                        
                        page_text = driver.find_element(By.TAG_NAME, "body").text.lower()
                        page_title = driver.title.lower()
                        
                        # Daftar jebakan Safe Browsing dan DNS Error
                        indikator_blokir = [
                            "deceptive site ahead", "situs menipu", "security error", "phishing", 
                            "this site can't be reached", "situs ini tidak dapat dijangkau", "err_name_not_resolved", "err_connection", 
                            "account has been suspended", "account suspended", "suspended page", 
                            "404 not found", "page not found"
                        ]
                        
                        terblokir = any(indikator in page_text or indikator in page_title for indikator in indikator_blokir)
                        
                        if terblokir:
                            hasil_scanner = "detected"
                            catatan = "Diblokir oleh Browser/Server (Safe Browsing / Dead Link)."
                            print("   🔴 Hasil Akhir: DIBLOKIR (Browser Warning / Web Mati)")
                        else:
                            hasil_scanner = "safe"
                            catatan = "Tembus total. Halaman termuat dengan sukses."
                            print("   🟢 Hasil Akhir: AMAN (Tembus Total)")
                            
                    except Exception as e:
                        # Kalau loading kelamaan atau DNS-nya ancur sampai Selenium gak bisa baca, kita anggap diblokir
                        hasil_scanner = "detected"
                        catatan = "Gagal memuat halaman (Timeout / Server Down)."
                        print("   🔴 Hasil Akhir: DIBLOKIR (Web tidak bisa diakses)")
                        
                    # Simpan hasil akhir ke Excel (Hanya 1 kali per gambar!)
                    logger.log_result(url_asli, label, kategori, hasil_scanner, catatan)
                    
                else:
                    hasil_scanner = "detected"
                    catatan = "Diblokir. Lens menetralisir link (Implicit Drop)."
                    print("   🔴 Hasil Lens: DIBLOKIR / TIDAK TERBACA")
                    logger.log_result(url_asli, label, kategori, hasil_scanner, catatan)
                
            except Exception as e:
                pesan_error = str(e).split('\n')[0]
                if "Timeout" in pesan_error or "timeout" in str(type(e)).lower():
                    pesan_error = "Timeout Render Lens"
                print(f"   🔴 Hasil Lens: DIBLOKIR (Fail-Safe: {pesan_error})")
                logger.log_result(url_asli, label, kategori, "detected", f"Diblokir (Error Ekstraksi: {pesan_error})")
                
            # Kita kembalikan timeout ke normal (300 detik) untuk Google Images berikutnya
            driver.set_page_load_timeout(300)
            jeda_manusia(2.0, 3.0) 
            
    except KeyboardInterrupt:
        print("\n🛑 Dihentikan paksa oleh pengguna.")
    finally:
        print("\n🏁 Pengujian selesai. Menutup browser...")
        driver.quit()

jalankan_robot()